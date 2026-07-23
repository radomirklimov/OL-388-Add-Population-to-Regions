import csv
from openpyxl import load_workbook
import shutil
import pandas as pd
import logging
import pandas as pd

proposal_file = "location_proposals.csv"
german_file = "population_germany.xlsx"
german_1_file = "population_germany_1.txt"
austria_file = "population_austria.csv"
austria_1_file = "population_austria_1.ods"
switzerland_file = "population_switzerland.csv"
switzerland_1_file = "population_switzerland_1.csv"

# Load german population
gp = load_workbook(german_file, read_only=True, data_only=True)
gpa = gp.active

# Build city -> population lookup for all cities in all countries
reference_population = {
    "DE": {},
    "AT": {},
    "CH": {}
}

# germany
for row in gpa.iter_rows(values_only=True):
    if row[6] is None or row[22] is None:
        continue

    name = str(row[6]).strip()
    level = str(row[1]).strip()

    if level in ("5", "6"):
        reference_population["DE"][name] = int(row[22])
gp.close()

# germany 1
with open(german_1_file, newline="", encoding="utf-8") as f:
    reader = csv.reader(f, delimiter="\t")

    for row in reader:
        name = row[1].strip()          # name
        population = row[14].strip()   # population

        if not population or population == "0":
            continue

        if reference_population["DE"].get(name) in (None, 0):
            reference_population["DE"][name] = int(population)

# austria
with open(austria_file, newline="", encoding="utf-8") as f:
    for line in f:
        if line.strip().startswith("ID;"):
            header = line
            break

    reader = csv.reader(f, delimiter=";")

    # Read the data
    for row in reader:
        name = row[1]
        population = int(row[2])

        reference_population["AT"][name] = population

# austria 1
df = pd.read_excel(austria_1_file, engine="odf", header=1)
# Remove leading/trailing whitespace from column names
df.columns = df.columns.str.strip()

for _, row in df.iterrows():
    # Skip rows with missing data
    if pd.isna(row["Ortschaftsname"]) or pd.isna(row["Bevölkerungam 01.01.2026"]) or row["Bevölkerungam 01.01.2026"] == 0:
        continue

    name = str(row["Ortschaftsname"]).strip()
    # Convert "10.492" -> 10492
    population = int(str(row["Bevölkerungam 01.01.2026"]).replace(".", ""))

    if reference_population["AT"].get(name) is None or reference_population["AT"].get(name) == 0:
        reference_population["AT"][name] = population

# switzerland
with open(switzerland_file, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f, delimiter=";")

    # Read the data
    for row in reader:
        name = row["GEO_NAME"]
        population = int(row["VALUE"])

        reference_population["CH"][name] = population

# switzerland 1
with open(switzerland_1_file, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)

    # Read the data
    for row in reader:
        name = row["Schweizer Städte"].strip()
        population = row["OBS_VALUE"].strip()

        # Skip missing values
        if not population:
            continue

        # Skip decimal values (e.g. "18.4")
        if "." in population:
            continue

        if reference_population["CH"].get(name) in (None, 0):
            reference_population["CH"][name] = int(population)

# Read proposal file
population_lookup = {}
non_population_cities = []
seen = set()
allCities = 0

with open(proposal_file, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)

    for row in reader:
        placeString = row["places"]

        if placeString in seen:
            continue

        places = [p.strip() for p in placeString.split(";")]
        allCities += len(places)

        total_population = 0

        for p in places:
            population = reference_population[row["country"]].get(p)

            if population is None or population == 0:
                non_population_cities.append(f"{row['country']}: {p}")
                print(f"WARN: no population of {p} found")
                continue

            seen.add(placeString)
            total_population += population

        if total_population != 0:
            population_lookup[placeString] = total_population


for city, population in population_lookup.items():
    print(f"{city}: {population}")

# ADD POPULATION COLUMN AND SAVE ALL FOUND POPULATIONS TO A NEW CSV FILE
# Read the CSV
lp = pd.read_csv("location_proposals.csv")

# Make a copy
lp_copy = lp.copy()

# Create the new column by looking up the population
lp_copy["population"] = lp_copy["places"].map(population_lookup)

# Optional: replace missing values with 0
lp_copy["population"] = lp_copy["population"].fillna(0).astype(int)

# Save the updated copy
lp_copy.to_csv("location_proposals_with_population.csv", index=False)

# TODO: collect all cities(independent of row, stand-alone city) for later research
logging.basicConfig(
    filename="non_population_cities.log",
    level=logging.INFO,
    format="%(message)s"
)

for city in non_population_cities:
    logging.info(f"City without population: {city}")
logging.info("")
logging.info(f"Total cities without population value: {len(non_population_cities)} from {allCities} total")


print()
print(f"Total cities without population value: {len(non_population_cities)} from {allCities} total")










# import csv
# from openpyxl import load_workbook

# proposal_file = "location_proposals.csv"
# reference_file = "reference_file.xlsx"

# # Load Excel workbook
# wb = load_workbook(reference_file, read_only=True, data_only=True)
# ws = wb.active

# # Build lookup: city name -> population
# population_lookup = dict()
# seen = set()
# nullPopulation = 0 

# def find_population(city_name):
#     for row in ws.iter_rows(values_only=True):
#         if row[6] is None or row[22] is None:
#             continue

#         name = str(row[6]).strip()
#         level = str(row[1]).strip()

#         if name == city_name and level in ("5", "6"):
#             return int(row[22])

#     return None  # not found

# # Read proposal file
# with open(proposal_file, newline="", encoding="utf-8") as f:
#     reader = csv.DictReader(f)

#     for row in reader:
#         placeString = row["places"]
#         if placeString in seen:
#             continue
#         places = [p.strip() for p in placeString.split(";")]

#         total_population = 0
#         for p in places:
#             population = find_population(p)

#             if population is None:
#                 print(f"WARN: no population of {p} found")
#                 continue

#             if population == 0:
#                 nullPopulation += 1

#             total_population += population

#         population_lookup[placeString] = total_population
#         seen.add(placeString)

# print(f"Number of cities whose population couldn't be found: {nullPopulation}")
# for city, population in population_lookup.items():
#     print(f"{city}: {population}")
# wb.close()








# # file_path = "location_proposals.csv"
# # base_url = "https://www.gemeindeverzeichnis.de/api/gemeinden"
# base_url1 = "http://api.geonames.org/searchJSON?q="
# base_url2 = "&country=DE&username=radomyrklymov"


# seen = set()
# last_request_time = 0

# nullPopulation = 0

# with open(file_path, mode="r", encoding="utf-8") as file:
#     reader = csv.DictReader(file)

#     for row in reader:
#         place = row["places"]

#         # Split on ';' if there are multiple places
#         places = [p.strip() for p in place.split(";")]

#         total_population = 0
#         for p in places:
#             # Cool-down: max 60 requests/minute
#             elapsed = time.time() - last_request_time
#             if elapsed < 1:
#                 time.sleep(1 - elapsed)

#             response = requests.get(base_url1 + p + base_url2)
#             response.raise_for_status()

#             data = response.json()

#             # total_population += data["results"][0]["population_total"]

#             if not data["geonames"]:
#                 print(f"WARN: no population of {p} found")
#                 continue

#             if data["geonames"][0]["population"] == 0:
#                 nullPopulation += 1
#                 print(nullPopulation)

#             total_population += data["geonames"][0]["population"]

#             # Respect rate limit
#             time.sleep(1)

#         key = (place, total_population)
#         if key in seen:
#             continue

#         seen.add(key)
#         print(f"{place}: {total_population}") 
